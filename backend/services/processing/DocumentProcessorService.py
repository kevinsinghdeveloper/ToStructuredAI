"""Document text extraction and chunking service."""
from io import BytesIO
from typing import List, Dict, Any
from abstractions.IServiceManagerBase import IServiceManagerBase


class DocumentProcessorService(IServiceManagerBase):
    """Extracts text from various document types and splits into chunks."""

    def __init__(self, config: dict = None):
        super().__init__(config or {})
        self.chunk_size = (config or {}).get("chunk_size", 1000)
        self.chunk_overlap = (config or {}).get("chunk_overlap", 200)
        self.storage_service = (config or {}).get("storage_service")

    def configure(self, **kwargs) -> None:
        pass

    def _get_file_object(self, file_path: str):
        """Return file path or BytesIO for remote files."""
        if file_path.startswith(("s3://", "gs://")):
            if not self.storage_service:
                raise Exception("Storage service not configured for remote file access")
            content = self.storage_service.download_file(file_path)
            return BytesIO(content)
        return file_path

    def run_task(self, request: Dict[str, Any]) -> Any:
        task_type = request.get("task_type")
        if task_type == "extract":
            return self.extract_text(request.get("file_path"), request.get("mime_type"))
        elif task_type == "chunk":
            return self.chunk_text(request.get("text"))
        raise ValueError(f"Unknown task type: {task_type}")

    def extract_text(self, file_path: str, mime_type: str) -> str:
        """Extract text from a document based on MIME type."""
        extractors = {
            "application/pdf": self._extract_from_pdf,
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": self._extract_from_docx,
            "application/msword": self._extract_from_docx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": self._extract_from_excel,
            "application/vnd.ms-excel": self._extract_from_excel,
            "text/csv": self._extract_from_csv,
            "application/csv": self._extract_from_csv,
            "text/plain": self._extract_from_txt,
        }
        extractor = extractors.get(mime_type)
        if not extractor:
            raise ValueError(f"Unsupported file type: {mime_type}")
        return extractor(file_path)

    def _extract_from_pdf(self, file_path: str) -> str:
        from pypdf import PdfReader
        reader = PdfReader(self._get_file_object(file_path))
        return "\n".join(page.extract_text() or "" for page in reader.pages).strip()

    def _extract_from_docx(self, file_path: str) -> str:
        from docx import Document
        doc = Document(self._get_file_object(file_path))
        return "\n".join(p.text for p in doc.paragraphs).strip()

    def _extract_from_excel(self, file_path: str) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(self._get_file_object(file_path))
        parts = []
        for sheet in wb.worksheets:
            parts.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                parts.append("\t".join(str(c) if c is not None else "" for c in row))
            parts.append("")
        return "\n".join(parts).strip()

    def _extract_from_csv(self, file_path: str) -> str:
        import csv
        from io import StringIO
        file_obj = self._get_file_object(file_path)
        if isinstance(file_obj, BytesIO):
            content = file_obj.read().decode("utf-8")
        else:
            with open(file_obj, "r", encoding="utf-8") as f:
                content = f.read()
        reader = csv.reader(StringIO(content))
        return "\n".join("\t".join(row) for row in reader).strip()

    def _extract_from_txt(self, file_path: str) -> str:
        file_obj = self._get_file_object(file_path)
        if isinstance(file_obj, BytesIO):
            return file_obj.read().decode("utf-8").strip()
        with open(file_obj, "r", encoding="utf-8") as f:
            return f.read().strip()

    def chunk_text(self, text: str) -> List[str]:
        """Split text into overlapping chunks, breaking at sentence boundaries."""
        chunks = []
        start = 0
        text_length = len(text)

        while start < text_length:
            end = start + self.chunk_size
            chunk = text[start:end]

            if end < text_length:
                last_period = chunk.rfind(".")
                last_newline = chunk.rfind("\n")
                break_point = max(last_period, last_newline)
                if break_point > self.chunk_size * 0.5:
                    chunk = chunk[: break_point + 1]
                    end = start + break_point + 1

            chunks.append(chunk.strip())
            start = end - self.chunk_overlap

        return chunks
