"""Export service - CSV, XLSX, PDF report generation with professional formatting."""
import io
import csv
from abstractions.IServiceManagerBase import IServiceManagerBase


EXPORT_COLUMNS = ["Date", "User", "Client", "Project", "Task", "Description",
                  "Duration (min)", "Hours", "Billable", "Hourly Rate", "Amount"]

# Column widths for PDF (landscape letter = ~720pt usable)
PDF_COL_WIDTHS = [58, 62, 62, 68, 58, 120, 42, 42, 42, 44, 50]

INVOICE_COL_WIDTHS = [58, 65, 65, 155, 50, 55, 58]

BRAND_COLOR = '#7b6df6'
BRAND_COLOR_LIGHT = '#ede9fe'
ALT_ROW_COLOR = '#f8f9fa'


def _row(e):
    return [
        e.get("date", ""),
        e.get("user_name", ""),
        e.get("client_name", ""),
        e.get("project_name", ""),
        e.get("task_name", ""),
        e.get("description", ""),
        e.get("duration_minutes", 0),
        e.get("hours", 0),
        "Yes" if e.get("is_billable") else "No",
        e.get("hourly_rate", ""),
        e.get("amount", ""),
    ]


class ExportService(IServiceManagerBase):
    def initialize(self):
        pass

    def export_entries(self, entries: list, fmt: str = "csv", metadata: dict = None) -> bytes:
        meta = metadata or {}
        if fmt == "csv":
            return self._to_csv(entries, meta)
        elif fmt == "xlsx":
            return self._to_xlsx(entries, meta)
        elif fmt == "pdf":
            return self._to_pdf(entries, meta)
        return self._to_csv(entries, meta)

    # ─── CSV ────────────────────────────────────────────────────────────

    def _to_csv(self, entries, meta=None):
        meta = meta or {}
        output = io.StringIO()
        writer = csv.writer(output)
        # Metadata header rows
        if meta.get("org_name"):
            writer.writerow([meta["org_name"]])
        if meta.get("report_title"):
            writer.writerow([meta["report_title"]])
        if meta.get("start_date") and meta.get("end_date"):
            writer.writerow([f"Period: {meta['start_date']} to {meta['end_date']}"])
        if any(meta.get(k) for k in ["org_name", "report_title", "start_date"]):
            writer.writerow([])  # blank separator
        writer.writerow(EXPORT_COLUMNS)
        for e in entries:
            writer.writerow(_row(e))
        # Summary row
        total_hours = sum(e.get("hours", 0) for e in entries)
        total_amount = sum(e.get("amount", 0) for e in entries)
        writer.writerow([])
        writer.writerow(["", "", "", "", "", "TOTAL", "",
                         f"{total_hours:.2f}", "", "", f"${total_amount:.2f}"])
        return output.getvalue().encode("utf-8")

    # ─── XLSX ───────────────────────────────────────────────────────────

    def _to_xlsx(self, entries, meta=None):
        meta = meta or {}
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = meta.get("report_title", "Time Entries")

            header_fill = PatternFill(start_color="7B6DF6", end_color="7B6DF6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            title_font = Font(bold=True, size=14, color="333333")
            subtitle_font = Font(bold=True, size=11, color="666666")
            period_font = Font(size=10, color="888888")
            total_font = Font(bold=True, size=10)
            thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))

            row_num = 1

            # Metadata rows
            if meta.get("org_name"):
                ws.cell(row=row_num, column=1, value=meta["org_name"]).font = title_font
                row_num += 1
            if meta.get("report_title"):
                ws.cell(row=row_num, column=1, value=meta["report_title"]).font = subtitle_font
                row_num += 1
            if meta.get("start_date") and meta.get("end_date"):
                ws.cell(row=row_num, column=1,
                        value=f"Period: {meta['start_date']} to {meta['end_date']}").font = period_font
                row_num += 1
            if row_num > 1:
                row_num += 1  # blank row

            # Column headers
            header_row = row_num
            for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            row_num += 1

            # Data rows
            for e in entries:
                row_data = _row(e)
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = thin_border
                    # Right-align and format number columns
                    if col_idx in (7, 8, 10, 11):
                        cell.alignment = Alignment(horizontal="right")
                        if col_idx in (10, 11) and isinstance(val, (int, float)):
                            cell.number_format = '$#,##0.00'
                        elif col_idx == 8 and isinstance(val, (int, float)):
                            cell.number_format = '#,##0.00'
                row_num += 1

            # Total row
            row_num += 1  # blank row
            total_hours = sum(e.get("hours", 0) for e in entries)
            total_amount = sum(e.get("amount", 0) for e in entries)
            ws.cell(row=row_num, column=6, value="TOTAL").font = total_font
            hrs_cell = ws.cell(row=row_num, column=8, value=total_hours)
            hrs_cell.font = total_font
            hrs_cell.number_format = '#,##0.00'
            hrs_cell.alignment = Alignment(horizontal="right")
            amt_cell = ws.cell(row=row_num, column=11, value=total_amount)
            amt_cell.font = total_font
            amt_cell.number_format = '$#,##0.00'
            amt_cell.alignment = Alignment(horizontal="right")

            # Auto column widths
            for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
                max_len = len(str(EXPORT_COLUMNS[col_idx - 1]))
                for row in ws.iter_rows(min_row=header_row, max_row=row_num,
                                        min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

            # Freeze header row
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        except ImportError:
            return self._to_csv(entries, meta)

    # ─── PDF ────────────────────────────────────────────────────────────

    def _to_pdf(self, entries, meta=None):
        meta = meta or {}
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.units import inch
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(letter),
                                    topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                                    leftMargin=0.5 * inch, rightMargin=0.5 * inch)
            styles = getSampleStyleSheet()

            brand = colors.HexColor(BRAND_COLOR)
            title_style = ParagraphStyle('RptTitle', parent=styles['Title'],
                                         fontSize=18, textColor=brand, spaceAfter=2)
            subtitle_style = ParagraphStyle('RptSub', parent=styles['Normal'],
                                            fontSize=10, textColor=colors.HexColor('#666666'),
                                            spaceAfter=12)

            elements = []

            # Header
            org = meta.get("org_name", "")
            title = meta.get("report_title", "Time Entry Report")
            if org:
                elements.append(Paragraph(org, title_style))
            elements.append(Paragraph(title, ParagraphStyle(
                'RptTitle2', parent=styles['Heading2'], fontSize=13, spaceAfter=2)))
            if meta.get("start_date") and meta.get("end_date"):
                elements.append(Paragraph(
                    f"Period: {meta['start_date']} to {meta['end_date']}", subtitle_style))
            else:
                elements.append(Spacer(1, 12))

            # Summary stats
            total_hours = sum(e.get("hours", 0) for e in entries)
            billable_hours = sum(e.get("hours", 0) for e in entries if e.get("is_billable"))
            total_amount = sum(e.get("amount", 0) for e in entries)
            summary_text = (f"<b>Total Hours:</b> {total_hours:.1f}  |  "
                            f"<b>Billable:</b> {billable_hours:.1f}  |  "
                            f"<b>Entries:</b> {len(entries)}  |  "
                            f"<b>Total Amount:</b> ${total_amount:,.2f}")
            elements.append(Paragraph(summary_text, ParagraphStyle(
                'RptStats', parent=styles['Normal'], fontSize=9,
                textColor=colors.HexColor('#444444'), spaceAfter=14)))

            # Table data
            data = [EXPORT_COLUMNS]
            for e in entries:
                row = _row(e)
                # Format numbers
                row[7] = f"{row[7]:.2f}" if isinstance(row[7], (int, float)) else str(row[7])
                row[10] = f"${row[10]:.2f}" if isinstance(row[10], (int, float)) else str(row[10])
                row[9] = f"${row[9]:.2f}" if isinstance(row[9], (int, float)) else str(row[9])
                data.append([str(v) for v in row])

            # Total row
            data.append(["", "", "", "", "", "TOTAL", "",
                         f"{total_hours:.2f}", "", "", f"${total_amount:,.2f}"])

            table = Table(data, colWidths=PDF_COL_WIDTHS, repeatRows=1)

            # Build style
            style_cmds = [
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), brand),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7.5),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                # Alignment
                ('ALIGN', (6, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (5, 0), 'LEFT'),
                # Grid
                ('LINEBELOW', (0, 0), (-1, 0), 1, brand),
                ('LINEBELOW', (0, -2), (-1, -2), 0.5, colors.HexColor('#cccccc')),
                # Total row
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#333333')),
                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]

            # Alternating row colors
            for i in range(1, len(data) - 1):
                if i % 2 == 0:
                    style_cmds.append(
                        ('BACKGROUND', (0, i), (-1, i), colors.HexColor(ALT_ROW_COLOR)))

            table.setStyle(TableStyle(style_cmds))
            elements.append(table)

            def _footer(canvas, doc_obj):
                canvas.saveState()
                canvas.setFont('Helvetica', 7)
                canvas.setFillColor(colors.HexColor('#999999'))
                canvas.drawRightString(doc_obj.pagesize[0] - 36, 28,
                                       f"Page {doc_obj.page}")
                canvas.drawString(36, 28, "Generated by Zerve My Time")
                canvas.restoreState()

            doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
            return output.getvalue()
        except ImportError:
            return self._to_csv(entries, meta)

    # ─── Invoice PDF ────────────────────────────────────────────────────

    def generate_invoice_pdf(self, invoice_data: dict) -> bytes:
        """Generate a professionally formatted invoice PDF."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import inch
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter,
                                    topMargin=0.6 * inch, bottomMargin=0.75 * inch,
                                    leftMargin=0.75 * inch, rightMargin=0.75 * inch)
            styles = getSampleStyleSheet()

            brand = colors.HexColor(BRAND_COLOR)
            brand_light = colors.HexColor(BRAND_COLOR_LIGHT)

            title_style = ParagraphStyle('InvTitle', parent=styles['Title'],
                                         fontSize=28, textColor=brand,
                                         spaceAfter=2, fontName='Helvetica-Bold')
            label_style = ParagraphStyle('InvLabel', parent=styles['Normal'],
                                         fontSize=8, textColor=colors.HexColor('#888888'),
                                         spaceAfter=1)
            value_style = ParagraphStyle('InvValue', parent=styles['Normal'],
                                         fontSize=10, spaceAfter=2)
            bold_value = ParagraphStyle('InvBold', parent=value_style,
                                        fontName='Helvetica-Bold')

            elements = []

            # ── Title + invoice number + date ──
            elements.append(Paragraph("INVOICE", title_style))

            inv_num = invoice_data.get("invoiceNumber", "")
            inv_date = invoice_data.get("date", "")
            meta_data = [[
                Paragraph(f"<b>Invoice #:</b> {inv_num}", value_style),
                Paragraph(f"<b>Date:</b> {inv_date}", value_style),
            ]]
            meta_table = Table(meta_data, colWidths=[250, 250])
            meta_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(meta_table)
            elements.append(Spacer(1, 16))

            # ── From / To (side by side) ──
            from_name = invoice_data.get("from", {}).get("name", "")
            to_info = invoice_data.get("to", {})

            from_parts = [Paragraph("FROM", label_style)]
            if from_name:
                from_parts.append(Paragraph(f"<b>{from_name}</b>", bold_value))

            to_parts = [Paragraph("BILL TO", label_style)]
            if to_info.get("name"):
                to_parts.append(Paragraph(f"<b>{to_info['name']}</b>", bold_value))
            if to_info.get("contactName"):
                to_parts.append(Paragraph(f"Attn: {to_info['contactName']}", value_style))
            if to_info.get("contactEmail"):
                to_parts.append(Paragraph(to_info["contactEmail"], value_style))
            if to_info.get("address"):
                to_parts.append(Paragraph(to_info["address"], value_style))

            # Combine from/to into cells
            from_cell = []
            for p in from_parts:
                from_cell.append(p)
            to_cell = []
            for p in to_parts:
                to_cell.append(p)

            addr_data = [[from_cell, to_cell]]
            addr_table = Table(addr_data, colWidths=[250, 250])
            addr_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (0, 0), (-1, -1), brand_light),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(addr_table)
            elements.append(Spacer(1, 12))

            # ── Project & Period ──
            period = invoice_data.get("period", {})
            elements.append(Paragraph(
                f"<b>Project:</b> {invoice_data.get('project', '')}", value_style))
            elements.append(Paragraph(
                f"<b>Period:</b> {period.get('start', '')} to {period.get('end', '')}",
                value_style))
            elements.append(Spacer(1, 14))

            # ── Line items table ──
            headers = ["Date", "User", "Task", "Description", "Hours", "Rate", "Amount"]
            data = [headers]
            for item in invoice_data.get("lineItems", []):
                desc = str(item.get("description", ""))
                if len(desc) > 80:
                    desc = desc[:77] + "..."
                data.append([
                    str(item.get("date", "")),
                    str(item.get("user", "")),
                    str(item.get("task", "")),
                    desc,
                    f"{item.get('hours', 0):.2f}",
                    f"${item.get('rate', 0):.2f}",
                    f"${item.get('amount', 0):.2f}",
                ])

            table = Table(data, colWidths=INVOICE_COL_WIDTHS, repeatRows=1)

            style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), brand),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('LINEBELOW', (0, 0), (-1, 0), 1, brand),
            ]
            # Alternating rows
            for i in range(1, len(data)):
                if i % 2 == 0:
                    style_cmds.append(
                        ('BACKGROUND', (0, i), (-1, i), colors.HexColor(ALT_ROW_COLOR)))
            # Bottom border on last data row
            style_cmds.append(('LINEBELOW', (0, -1), (-1, -1), 0.5,
                               colors.HexColor('#cccccc')))

            table.setStyle(TableStyle(style_cmds))
            elements.append(table)
            elements.append(Spacer(1, 8))

            # ── Totals section (right-aligned) ──
            subtotal = invoice_data.get("totalAmount", 0)
            tax_rate = invoice_data.get("taxRate", 0)
            tax_amount = round(subtotal * tax_rate / 100, 2) if tax_rate else 0
            grand_total = subtotal + tax_amount

            totals_data = [
                ["", f"Subtotal: {invoice_data.get('totalHours', 0):.2f} hrs",
                 f"${subtotal:,.2f}"],
            ]
            if tax_rate:
                totals_data.append(["", f"Tax ({tax_rate}%):", f"${tax_amount:,.2f}"])
            totals_data.append(["", "TOTAL:", f"${grand_total:,.2f}"])

            totals_table = Table(totals_data, colWidths=[310, 110, 86])
            total_row_idx = len(totals_data) - 1
            totals_table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('FONTNAME', (1, total_row_idx), (-1, total_row_idx), 'Helvetica-Bold'),
                ('FONTSIZE', (1, total_row_idx), (-1, total_row_idx), 12),
                ('TEXTCOLOR', (1, total_row_idx), (-1, total_row_idx), brand),
                ('LINEABOVE', (1, total_row_idx), (-1, total_row_idx), 1,
                 colors.HexColor('#333333')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(totals_table)
            elements.append(Spacer(1, 24))

            # ── Payment terms ──
            elements.append(Paragraph("<b>Payment Terms:</b> Net 30", ParagraphStyle(
                'Terms', parent=styles['Normal'], fontSize=9,
                textColor=colors.HexColor('#666666'))))
            elements.append(Paragraph(
                f"<b>Currency:</b> {invoice_data.get('currency', 'USD')}",
                ParagraphStyle('Currency', parent=styles['Normal'], fontSize=9,
                               textColor=colors.HexColor('#666666'))))

            # ── Footer ──
            def _footer(canvas, doc_obj):
                canvas.saveState()
                canvas.setFont('Helvetica', 7)
                canvas.setFillColor(colors.HexColor('#999999'))
                canvas.drawString(54, 36, "Generated by Zerve My Time")
                canvas.drawRightString(doc_obj.pagesize[0] - 54, 36,
                                       f"Page {doc_obj.page}")
                canvas.restoreState()

            doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
            return output.getvalue()
        except ImportError:
            return self._generate_invoice_csv(invoice_data)

    def _generate_invoice_csv(self, invoice_data: dict) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Invoice {invoice_data.get('invoiceNumber', '')}"])
        writer.writerow([f"Date: {invoice_data.get('date', '')}"])
        writer.writerow([f"Project: {invoice_data.get('project', '')}"])
        writer.writerow([])
        writer.writerow(["Date", "User", "Task", "Description", "Hours", "Rate", "Amount"])
        for item in invoice_data.get("lineItems", []):
            writer.writerow([item.get("date"), item.get("user"), item.get("task"),
                             item.get("description"), item.get("hours"),
                             item.get("rate"), item.get("amount")])
        writer.writerow([])
        writer.writerow(["", "", "", "TOTAL", invoice_data.get("totalHours"),
                         "", invoice_data.get("totalAmount")])
        return output.getvalue().encode("utf-8")
